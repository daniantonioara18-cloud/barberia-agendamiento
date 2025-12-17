# Barberia/views.py
from datetime import datetime, timedelta, date

import csv

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse, HttpResponseBadRequest
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.urls import reverse

from django.contrib import messages
from django.contrib.auth import authenticate, login as dj_login
from django.contrib.auth.decorators import login_required, user_passes_test

from django.template.loader import render_to_string
from xhtml2pdf import pisa

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from django.core.exceptions import ValidationError

from .models import Horas, Horario, Usuario, Dias, Tipo_servicio, DiaCerrado
from django.views.decorators.http import require_GET

# =========================
# Constantes / helpers base
# =========================

DIAS_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
ORDEN_DIAS_ATENCION = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]


def _solo_staff(u):
    return u.is_authenticated and u.is_staff


def semana_actual():
    """
    Chips de semana: lunes..domingo de la semana actual.
    Retorna (lista, hoy)
    lista = [{date, value, label}, ...]
    """
    hoy = timezone.localdate()
    lunes = hoy - timedelta(days=hoy.weekday())  # 0=lunes
    dias = []
    for i in range(7):
        d = lunes + timedelta(days=i)
        label = f"{DIAS_ES[d.weekday()]} — {d.day:02d}"
        dias.append({
            "date": d,
            "value": d.isoformat(),
            "label": label,
        })
    return dias, hoy


def _duracion_minutos(servicio: Tipo_servicio) -> int:
    nombre = (getattr(servicio, "nombre", "") or "").lower()
    return 60 if "perfil" in nombre else 30


def _nombre_dia(fecha: date) -> str:
    return DIAS_ES[fecha.weekday()]


def _ventana_reservable():
    """
    Ventana permitida:
    - Lun..Sáb: hoy .. sábado de esta semana
    - Dom:     lunes próximo .. sábado próximo
    """
    hoy = timezone.localdate()
    wd = hoy.weekday()  # 0=lun .. 6=dom
    if wd == 6:  # domingo
        lunes = hoy + timedelta(days=1)
        sabado = lunes + timedelta(days=5)
        fecha_min = lunes
    else:
        lunes = hoy - timedelta(days=wd)
        sabado = lunes + timedelta(days=5)
        fecha_min = hoy
    return fecha_min, sabado


# =========================
# PANEL: Calendario + Export + Stats
# =========================

@login_required
@user_passes_test(_solo_staff)
def panel_calendario(request):
    servicios = Tipo_servicio.objects.all().order_by("nombre")
    hoy = timezone.localdate()

    if request.method == 'POST':
        if 'cerrar_hoy' in request.POST:
            DiaCerrado.objects.get_or_create(fecha=hoy)
            messages.success(request, 'Has marcado HOY como día cerrado')
            return redirect('panel_calendario')

        if 'abrir_hoy' in request.POST:
            DiaCerrado.objects.filter(fecha=hoy).delete()
            messages.success(request, 'Has quitado el cierre de HOY')
            return redirect('panel_calendario')

    hoy_cerrado = DiaCerrado.objects.filter(fecha=hoy).exists()
    return render(request, "panel/calendario.html", {
        "servicios": servicios,
        "hoy": hoy,
        "hoy_cerrado": hoy_cerrado,
    })


@login_required
@user_passes_test(_solo_staff)
def panel_api_events(request):
    """
    Endpoint FullCalendar.
    GET ?start=YYYY-MM-DD&end=YYYY-MM-DD (end exclusivo)
    Opcional: servicio=ID, estado=P/A/C
    """
    start = request.GET.get("start")
    end = request.GET.get("end")
    servicio = request.GET.get("servicio") or ""
    estado = request.GET.get("estado") or ""

    try:
        start_date = datetime.fromisoformat(start[:10]).date() if start else None
        end_date = datetime.fromisoformat(end[:10]).date() if end else None
    except Exception:
        start_date = end_date = None

    qs = (
        Horario.objects
        .select_related("usuario_horario", "hora_horario", "Tipo_servicio", "dia_horario")
        .order_by("fecha", "hora_horario__hora_Horas")
    )

    if start_date and end_date:
        qs = qs.filter(fecha__gte=start_date, fecha__lt=end_date)

    if servicio.isdigit():
        qs = qs.filter(Tipo_servicio_id=int(servicio))

    if estado in dict(Horario.ESTADOS).keys():
        qs = qs.filter(estado=estado)

    COLOR_ESTADO = {
        "P": "#f59e0b",
        "A": "#22c55e",
        "C": "#ef4444",
    }

    events = []
    for h in qs:
        if not h.fecha:
            continue

        hhmm = getattr(h.hora_horario, "hora_Horas", "09:00")
        try:
            hour, minute = map(int, hhmm.split(":")[:2])
        except Exception:
            hour, minute = 9, 0

        start_dt = datetime(h.fecha.year, h.fecha.month, h.fecha.day, hour, minute)
        mins = _duracion_minutos(h.Tipo_servicio)
        end_dt = start_dt + timedelta(minutes=mins)

        title = f"{h.usuario_horario.nombre} · {h.Tipo_servicio.nombre}"

        events.append({
            "id": h.id,
            "title": title,
            "start": start_dt.isoformat(),
            "end": end_dt.isoformat(),
            "color": COLOR_ESTADO.get(h.estado, "#94a3b8"),
            "extendedProps": {
                "estado": h.get_estado_display(),
                "servicio": h.Tipo_servicio.nombre,
                "hora": hhmm,
                "dia_semana": getattr(h.dia_horario, "dia_Dias", ""),
                "rut": getattr(h.usuario_horario, "rut", ""),
                "celular": getattr(h.usuario_horario, "celular", ""),
            }
        })

    return JsonResponse(events, safe=False)


@login_required
@user_passes_test(_solo_staff)
def panel_export_rango(request):
    start_str = request.GET.get("start")
    end_str = request.GET.get("end")  # end exclusivo
    servicio = request.GET.get("servicio")
    estado = request.GET.get("estado")

    if not (start_str and end_str):
        return HttpResponseBadRequest("Faltan start/end")

    start = parse_date(start_str)
    end = parse_date(end_str)
    if not (start and end):
        return HttpResponseBadRequest("Fechas inválidas")

    qs = (
        Horario.objects
        .select_related('usuario_horario', 'hora_horario', 'Tipo_servicio', 'dia_horario')
        .filter(fecha__gte=start, fecha__lt=end)
        .order_by('fecha', 'hora_horario__hora_Horas')
    )

    if servicio and servicio.isdigit():
        qs = qs.filter(Tipo_servicio_id=int(servicio))
    if estado in dict(Horario.ESTADOS).keys():
        qs = qs.filter(estado=estado)

    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="citas_{start:%Y%m%d}_{end:%Y%m%d}.csv"'

    w = csv.writer(resp)
    w.writerow(['ID', 'Fecha', 'Hora', 'Cliente', 'RUT', 'Teléfono', 'Servicio', 'Día', 'Estado'])

    estado_label = dict(Horario.ESTADOS)
    for c in qs:
        w.writerow([
            c.id,
            (c.fecha.isoformat() if c.fecha else ""),
            c.hora_horario.hora_Horas,
            c.usuario_horario.nombre,
            c.usuario_horario.rut,
            c.usuario_horario.celular,
            c.Tipo_servicio.nombre,
            c.dia_horario.dia_Dias,
            estado_label.get(c.estado, ''),
        ])
    return resp


@login_required
@user_passes_test(_solo_staff)
def panel_export_rango_excel(request):
    start_str = request.GET.get("start")
    end_str = request.GET.get("end")  # end exclusivo
    servicio = request.GET.get("servicio")
    estado = request.GET.get("estado")

    if not (start_str and end_str):
        return HttpResponseBadRequest("Faltan start/end")

    start = parse_date(start_str)
    end = parse_date(end_str)
    if not (start and end):
        return HttpResponseBadRequest("Fechas inválidas")

    qs = (
        Horario.objects
        .select_related('usuario_horario', 'hora_horario', 'Tipo_servicio', 'dia_horario')
        .filter(fecha__gte=start, fecha__lt=end)
        .order_by('fecha', 'hora_horario__hora_Horas')
    )

    if servicio and servicio.isdigit():
        qs = qs.filter(Tipo_servicio_id=int(servicio))
    if estado in dict(Horario.ESTADOS).keys():
        qs = qs.filter(estado=estado)

    wb = Workbook()
    ws = wb.active
    ws.title = "Citas"

    headers = ['ID', 'Fecha', 'Hora', 'Cliente', 'RUT', 'Teléfono', 'Servicio', 'Día', 'Estado']
    ws.append(headers)

    estado_label = dict(Horario.ESTADOS)
    for c in qs:
        ws.append([
            c.id,
            (c.fecha.strftime("%d-%m-%Y") if c.fecha else ""),
            c.hora_horario.hora_Horas,
            c.usuario_horario.nombre,
            c.usuario_horario.rut,
            c.usuario_horario.celular,
            c.Tipo_servicio.nombre,
            c.dia_horario.dia_Dias,
            estado_label.get(c.estado, ""),
        ])

    header_fill = PatternFill("solid", fgColor="111111")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, len(headers) + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 40)

    if ws.max_row >= 1:
        last_col = get_column_letter(len(headers))
        table_ref = f"A1:{last_col}{ws.max_row}"
        tab = Table(displayName="TablaCitas", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

    filename = f'citas_{start:%Y%m%d}_{end:%Y%m%d}.xlsx'
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


@login_required
@user_passes_test(_solo_staff)
def panel_api_stats(request):
    """
    GET ?start=YYYY-MM-DD&end=YYYY-MM-DD (end exclusivo)
    Opcional: estado=P/A/C  servicio=ID
    """
    start_str = request.GET.get("start")
    end_str = request.GET.get("end")
    estado = request.GET.get("estado") or ""
    servicio = request.GET.get("servicio") or ""

    if not (start_str and end_str):
        return JsonResponse({"labels": [], "values": []})

    start = parse_date(start_str)
    end = parse_date(end_str)
    if not (start and end):
        return JsonResponse({"labels": [], "values": []})

    qs = Horario.objects.select_related("Tipo_servicio").filter(
        fecha__gte=start,
        fecha__lt=end
    )

    if estado in dict(Horario.ESTADOS).keys():
        qs = qs.filter(estado=estado)

    if servicio.isdigit():
        qs = qs.filter(Tipo_servicio_id=int(servicio))

    data = (
        qs.values("Tipo_servicio__nombre")
        .annotate(total=Count("id"))
        .order_by("-total", "Tipo_servicio__nombre")
    )

    labels = [r["Tipo_servicio__nombre"] or "Sin servicio" for r in data]
    values = [r["total"] for r in data]
    return JsonResponse({"labels": labels, "values": values})


@login_required
@user_passes_test(_solo_staff)
def panel_api_canceladas(request):
    """
    Tabla canceladas debajo del calendario
    GET ?start=YYYY-MM-DD&end=YYYY-MM-DD (end exclusivo)
    Opcional: servicio=ID
    """
    start_str = request.GET.get("start")
    end_str = request.GET.get("end")
    servicio = request.GET.get("servicio") or ""

    start = parse_date(start_str) if start_str else None
    end = parse_date(end_str) if end_str else None
    if not (start and end):
        return JsonResponse({"rows": []})

    qs = (
        Horario.objects
        .select_related("usuario_horario", "hora_horario", "Tipo_servicio")
        .filter(fecha__gte=start, fecha__lt=end, estado="C")
        .order_by("fecha", "hora_horario__hora_Horas")
    )

    if servicio.isdigit():
        qs = qs.filter(Tipo_servicio_id=int(servicio))

    rows = [{
        "nombre": h.usuario_horario.nombre,
        "rut": h.usuario_horario.rut,
        "fecha": h.fecha.strftime("%d-%m-%Y") if h.fecha else "",
        "hora": h.hora_horario.hora_Horas,
        "servicio": h.Tipo_servicio.nombre,
    } for h in qs]

    return JsonResponse({"rows": rows})


# =========================
# PANEL: Tabla horarios + filtros
# =========================

def _filtar_citas(request):
    q = (request.GET.get('q') or '').strip()
    dia = request.GET.get('dia') or ''
    servicio = request.GET.get('servicio') or ''
    estado = request.GET.get('estado') or ''

    qs = (
        Horario.objects
        .select_related('usuario_horario', 'hora_horario', 'Tipo_servicio', 'dia_horario')
        .order_by('fecha', 'hora_horario__hora_Horas')
    )

    if q:
        qs = qs.filter(
            Q(usuario_horario__nombre__icontains=q) |
            Q(usuario_horario__rut__icontains=q)
        )
    if dia.isdigit():
        qs = qs.filter(dia_horario__id=int(dia))
    if servicio.isdigit():
        qs = qs.filter(Tipo_servicio__id=int(servicio))
    if estado in dict(Horario.ESTADOS).keys():
        qs = qs.filter(estado=estado)

    return qs, q, dia, servicio, estado


@login_required
@user_passes_test(_solo_staff)
def panel_horarios(request):
    qs, q, dia, servicio, estado = _filtar_citas(request)

    week_days, _ = semana_actual()

    fecha_q = (request.GET.get("fecha") or "").strip()
    if fecha_q:
        try:
            f = date.fromisoformat(fecha_q)
            qs = qs.filter(fecha=f)
        except ValueError:
            pass

    qs_activas = qs.filter(estado__in=["P", "A"])
    qs_cancel = qs.filter(estado="C")

    citas = Paginator(qs_activas, 10).get_page(request.GET.get('page'))
    canceladas = Paginator(qs_cancel, 10).get_page(request.GET.get('page_c'))

    dias = Dias.objects.all().order_by('id')
    servicios = Tipo_servicio.objects.all().order_by('nombre')

    return render(request, 'panel/horarios_list.html', {
        'citas': citas,
        'canceladas': canceladas,
        'dias': dias,
        'servicios': servicios,
        'q': q,
        'dia_sel': dia,
        'servicio_sel': servicio,
        'estado_sel': estado,
        'week_days': week_days,
        'selected_date': fecha_q,
    })


@login_required
@user_passes_test(_solo_staff)
def panel_export(request):
    qs, q, dia, servicio, estado = _filtar_citas(request)

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="citas_{timezone.now():%Y%m%d_%H%M}.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Cliente', 'RUT', 'Día', 'Hora', 'Servicio', 'Estado'])

    estado_label = dict(Horario.ESTADOS)
    for c in qs:
        writer.writerow([
            c.id,
            getattr(c.usuario_horario, 'nombre', str(c.usuario_horario)),
            getattr(c.usuario_horario, 'rut', ''),
            str(c.dia_horario),
            str(c.hora_horario),
            getattr(c.Tipo_servicio, 'nombre', str(c.Tipo_servicio)),
            estado_label.get(c.estado, '')
        ])
    return response


@login_required
@user_passes_test(_solo_staff)
def panel_set_estado(request, pk):
    if request.method != 'POST':
        return redirect(request.META.get('HTTP_REFERER', reverse('panel_horarios')))

    nuevo = request.POST.get('estado')
    validos = dict(Horario.ESTADOS).keys()
    if nuevo not in validos:
        messages.error(request, "Estado inválido.")
        return redirect(request.META.get('HTTP_REFERER', reverse('panel_horarios')))

    h = get_object_or_404(Horario, pk=pk)

    # Evita “revivir” una cancelada si ya existe una activa en esa misma fecha+hora
    if h.estado == "C" and nuevo in ["P", "A"]:
        conflicto = Horario.objects.filter(
            fecha=h.fecha,
            hora_horario=h.hora_horario,
            estado__in=["P", "A"]
        ).exclude(pk=h.pk).exists()

        if conflicto:
            messages.error(
                request,
                "No se puede reactivar: esa hora ya fue tomada por otra cita."
            )
            return redirect(request.META.get('HTTP_REFERER', reverse('panel_horarios')))

    h.estado = nuevo
    h.save(update_fields=["estado"])
    messages.success(request, "Estado actualizado.")
    return redirect(request.META.get('HTTP_REFERER', reverse('panel_horarios')))


def login_view(request):
    if request.method == 'POST':
        u = request.POST.get('username', '')
        p = request.POST.get('password', '')
        user = authenticate(request, username=u, password=p)
        if user is not None:
            dj_login(request, user)
            messages.success(request, "Bienvenido, " + user.get_username())
            return redirect('home')
        messages.error(request, "Usuario o contraseña incorrectos.")
    return render(request, 'panel/login.html')


# =========================
# Sitio público (index / consultas / agendar)
# =========================

def mostrarindex(request):
    return render(request, 'index.html')


def consultas(request):
    servicios = Tipo_servicio.objects.all().order_by('nombre')
    horario = [
        ('Lunes', '12:00 – 19:00'),
        ('Martes', '12:00 – 19:00'),
        ('Miércoles', '12:00 – 19:00'),
        ('Jueves', '12:00 – 19:00'),
        ('Viernes', '12:00 – 19:00'),
        ('Sábado', '12:00 – 19:30'),
        ('Domingo', 'Cerrado'),
    ]
    return render(request, 'consulta.html', {
        'servicios': servicios,
        'horario_atencion': horario,
    })


def mostrarSobreMI(request):
    return render(request, 'sobreMI.html')


def mostrarAgendamiento(request):
    """
    Render del formulario agendar.html
    """
    tipos = Tipo_servicio.objects.all()

    dias_db = set(Dias.objects.values_list("dia_Dias", flat=True))
    dias_dis = [d for d in ORDEN_DIAS_ATENCION if d in dias_db]

    hoy = timezone.localdate()
    hoy_nombre = DIAS_ES[hoy.weekday()]
    hoy_cerrado = DiaCerrado.objects.filter(fecha=hoy).exists()

    return render(request, "agendar.html", {
        "tipos_disponibles": tipos,
        "dias_disponibles": dias_dis,
        "hoy_nombre": hoy_nombre,
        "hoy_cerrado": hoy_cerrado,
    })


def AgendarCita(request):
    return render(request, "agendar.html", {"dias_disponibles": Dias.objects.all()})


# =========================
# API Slots + lógica de disponibilidad
# =========================

def _calcular_slots_disponibles(fecha: date, servicio_id: int):
    """
    Slots disponibles para UNA fecha.
    - Ocupadas: estado P o A (CANCELADA no bloquea)
    - Si fecha == HOY: NO mostrar horas que ya pasaron (sin margen)
    - Si servicio dura 60': necesita 2 bloques consecutivos libres
    """
    if _nombre_dia(fecha) == "Domingo":
        return []

    horas = list(Horas.objects.order_by("id").values_list("hora_Horas", flat=True))

    ocupadas = set(
        Horario.objects.filter(fecha=fecha, estado__in=["P", "A"])
        .values_list("hora_horario__hora_Horas", flat=True)
    )

    servicio = get_object_or_404(Tipo_servicio, id=servicio_id)
    es_60 = "perfil" in (servicio.nombre or "").lower()

    libres = [h for h in horas if h not in ocupadas]

    # Filtrar horas pasadas si es HOY
    hoy = timezone.localdate()
    if fecha == hoy:
        now = timezone.localtime(timezone.now())
        h_now, m_now = now.hour, now.minute

        def _es_futura(hhmm: str) -> bool:
            try:
                hh, mm = map(int, hhmm.split(":")[:2])
            except Exception:
                return False
            return (hh, mm) > (h_now, m_now)

        libres = [h for h in libres if _es_futura(h)]

    if not es_60:
        return libres

    # 60': pares consecutivos
    consecutivos = []
    libres_set = set(libres)
    for i in range(len(horas) - 1):
        h1 = horas[i]
        h2 = horas[i + 1]
        if h1 in libres_set and h2 in libres_set:
            consecutivos.append(h1)
    return consecutivos


def api_slots(request):
    """
    GET /api/slots?fecha=YYYY-MM-DD&servicio=ID
    Respuesta: {"slots": ["12:00","12:30", ...]}
    """
    fecha_str = request.GET.get("fecha")
    servicio_id = request.GET.get("servicio")
    if not fecha_str or not servicio_id:
        return JsonResponse({"slots": []})

    try:
        f = date.fromisoformat(fecha_str)
    except ValueError:
        return JsonResponse({"slots": []})

    if DiaCerrado.objects.filter(fecha=f).exists():
        return JsonResponse({"slots": []})

    fecha_min, fecha_max = _ventana_reservable()
    if f < fecha_min or f > fecha_max:
        return JsonResponse({"slots": []})

    if _nombre_dia(f) == "Domingo":
        return JsonResponse({"slots": []})

    try:
        servicio_id_int = int(servicio_id)
    except (TypeError, ValueError):
        return JsonResponse({"slots": []})

    slots = _calcular_slots_disponibles(f, servicio_id_int)
    return JsonResponse({"slots": slots})




@require_GET
def api_ocupadas(request):
    """
    GET /api/ocupadas?fecha=YYYY-MM-DD
    Devuelve las horas NO disponibles (ocupadas) para esa fecha.
    Considera ocupadas: estado P o A (Canceladas NO bloquean).
    """
    fecha_str = request.GET.get("fecha")
    if not fecha_str:
        return JsonResponse({"rows": []})

    try:
        f = date.fromisoformat(fecha_str)
    except ValueError:
        return JsonResponse({"rows": []})

    # si quieres que día cerrado muestre todo como "no disponible", podrías manejarlo aquí,
    # pero por ahora mostramos solo las ocupadas en BD.
    qs = (
        Horario.objects
        .select_related("hora_horario", "dia_horario")
        .filter(fecha=f, estado__in=["P", "A"])
        .order_by("hora_horario__hora_Horas")
    )

    rows = []
    for h in qs:
        rows.append({
            "dia": h.dia_horario.dia_Dias,
            "fecha": h.fecha.strftime("%d-%m-%Y") if h.fecha else "",
            "hora": h.hora_horario.hora_Horas,
        })

    return JsonResponse({"rows": rows})


# =========================
# Registrar agendamiento (POST)
# =========================

def RegistrarHorario(request):
    if request.method != "POST":
        return redirect("/Agendamiento/")

    name = request.POST.get("name")
    rutificador = request.POST.get("rut")
    celu = request.POST.get("telefono")
    servicio_id = request.POST.get("servicio_id")
    hora_str = request.POST.get("hora")
    fecha_str = request.POST.get("fecha")

    if not all([name, rutificador, celu, servicio_id, hora_str, fecha_str]):
        return HttpResponseBadRequest("Faltan datos.")

    try:
        f = date.fromisoformat(fecha_str)
    except ValueError:
        return HttpResponseBadRequest("Fecha inválida.")

    if DiaCerrado.objects.filter(fecha=f).exists():
        return HttpResponseBadRequest("Este día está cerrado. Selecciona otra fecha.")

    fecha_min, fecha_max = _ventana_reservable()
    if f < fecha_min or f > fecha_max:
        return HttpResponseBadRequest("Solo se puede agendar en la ventana permitida.")

    nombre_dia = _nombre_dia(f)
    if nombre_dia == "Domingo":
        return HttpResponseBadRequest("No se atiende los domingos.")

    dia_obj = get_object_or_404(Dias, dia_Dias=nombre_dia)
    hora_obj = get_object_or_404(Horas, hora_Horas=hora_str)

    try:
        servicio_id_int = int(servicio_id)
    except (TypeError, ValueError):
        return HttpResponseBadRequest("Servicio inválido.")

    slots_validos = _calcular_slots_disponibles(f, servicio_id_int)
    if hora_str not in slots_validos:
        return HttpResponseBadRequest("Esa hora no está disponible para el servicio elegido.")

    if Horario.objects.filter(fecha=f, hora_horario=hora_obj, estado__in=["P", "A"]).exists():
        return HttpResponseBadRequest("Esta hora ya está ocupada, selecciona otra.")

    try:
        user = Usuario.objects.create(
            nombre=name,
            celular=celu,
            rut=rutificador
        )
    except ValidationError as e:
        mensaje = e.message_dict.get("rut", ["El RUT ingresado no es válido."])[0]
        messages.error(request, mensaje)
        return redirect("/Agendamiento/")

    horario = Horario.objects.create(
        usuario_horario=user,
        dia_horario=dia_obj,
        hora_horario=hora_obj,
        Tipo_servicio_id=servicio_id_int,
        fecha=f,
        estado='P'
    )

    return render(request, "exito.html", {
        "nombre": user.nombre,
        "rut": user.rut,
        "fecha": f.strftime("%d-%m-%Y"),
        "hora": hora_obj.hora_Horas,
        "horario_id": horario.id,
    })


# =========================
# Legacy / compat
# =========================

def obtener_horas_disponibles(request, dia_id):
    horas_ocupadas = Horario.objects.filter(dia_horario_id=dia_id).values_list('hora_horario_id', flat=True)
    horas_disponibles = Horas.objects.exclude(id__in=horas_ocupadas)
    data = {'horas': [{'id': h.id, 'hora': h.hora_Horas} for h in horas_disponibles]}
    return JsonResponse(data)


# =========================
# PDF + Listado público
# =========================

def generar_comprobante(request, horario_id):
    horario = get_object_or_404(Horario, id=horario_id)
    contexto = {
        'nombre': horario.usuario_horario.nombre,
        'rut': horario.usuario_horario.rut,
        'fecha': (horario.fecha.strftime("%d-%m-%Y") if horario.fecha else ""),
        'hora': horario.hora_horario.hora_Horas,
    }
    html = render_to_string('comprobante.html', contexto)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="comprobante_{horario.id}.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Hubo un error al generar el PDF.')
    return response


def mostrarlistadoHora(request):
    week_days, hoy = semana_actual()
    fecha_q = request.GET.get("fecha")
    f = date.fromisoformat(fecha_q) if fecha_q else hoy

    citas = (
        Horario.objects
        .filter(fecha=f)
        .exclude(estado='C')
        .select_related('usuario_horario', 'hora_horario', 'Tipo_servicio', 'dia_horario')
        .order_by('hora_horario__hora_Horas')
    )

    selected_label = f"{_nombre_dia(f)} {f:%d-%m-%Y}"

    return render(request, "listadohora.html", {
        "citas": citas,
        "week_days": week_days,
        "selected_date": f,
        "selected_label": selected_label,
    })
